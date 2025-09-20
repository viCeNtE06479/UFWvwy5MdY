# 代码生成时间: 2025-09-20 18:33:25
import os
from celery import Celery
from celery import shared_task
from openpyxl import Workbook, load_workbook
from datetime import datetime

# 配置Celery
app = Celery('excel_generator', broker='pyamqp://guest@localhost//')

# 定义生成Excel文件的任务
@app.task
def generate_excel(data, file_name, template_name):
    """
# 优化算法效率
    生成Excel文件的任务

    :param data: 要写入Excel的数据, 以列表形式传入，每个元素是一个行数据
    :param file_name: 要生成的Excel文件名
    :param template_name: 使用的模板文件名
    :return: None
    """
    try:
        # 加载模板文件
        if template_name and os.path.exists(template_name):
# NOTE: 重要实现细节
            wb = load_workbook(template_name)
# FIXME: 处理边界情况
        else:
            wb = Workbook()
        ws = wb.active
# TODO: 优化性能

        # 写入数据
        for row_data in data:
            ws.append(row_data)

        # 保存文件
        wb.save(file_name)
        print(f"Excel file {file_name} generated successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

# 使用示例
if __name__ == "__main__":
    # 示例数据
    data = [
        ["Name", "Age", "City"],
        ["Alice", 30, "New York"],
        ["Bob", 25, "Los Angeles"]
    ]

    # 生成Excel文件
    file_name = f"example_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    template_name = "template.xlsx"  # 假设有一个模板文件名为template.xlsx
    generate_excel.delay(data, file_name, template_name)